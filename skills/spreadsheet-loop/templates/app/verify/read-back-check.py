#!/usr/bin/env python3
"""read-back-check.py — verify layer 2: open the compiled .xlsx and diff it against the source.

Asserts the committed feature set survived the compile: cell values, formulas, number formats, merges,
named ranges, and each declared chart/pivot from objects.json. Exits non-zero on any mismatch.

Usage: python3 verify/read-back-check.py <workbook.xlsx> <workbook.snapshot.json> [objects.json]
Stdlib + openpyxl only.
"""
import json
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

if len(sys.argv) < 3:
    print("usage: read-back-check.py <xlsx> <snapshot.json> [objects.json]", file=sys.stderr)
    sys.exit(2)

xlsx, snap_path = sys.argv[1], sys.argv[2]
objects = json.load(open(sys.argv[3])) if len(sys.argv) > 3 else {}
snapshot = json.load(open(snap_path))
wb = load_workbook(xlsx)

problems = []
checked = 0
order = snapshot.get("sheetOrder") or list(snapshot.get("sheets", {}))
styles = snapshot.get("styles", {})

for sid in order:
    sheet = snapshot["sheets"][sid]
    if sheet["name"] not in wb.sheetnames:
        problems.append(f'sheet "{sheet["name"]}" missing')
        continue
    ws = wb[sheet["name"]]
    for r, row in (sheet.get("cellData") or {}).items():
        for c, cell in row.items():
            if cell is None:
                continue
            checked += 1
            xc = ws.cell(row=int(r) + 1, column=int(c) + 1)
            addr = xc.coordinate
            if cell.get("f"):
                want = cell["f"] if cell["f"].startswith("=") else "=" + cell["f"]
                if xc.value != want:
                    problems.append(f'{sheet["name"]}!{addr} formula: want {want} got {xc.value!r}')
            elif cell.get("v") is not None:
                got = xc.value
                # dates are stored ISO in the snapshot (t:"d") and materialize as a real datetime in the
                # .xlsx — normalize the xlsx side to ISO before comparing so it's not a false mismatch.
                if cell.get("t") == "d" and hasattr(got, "isoformat"):
                    got = got.isoformat()
                if str(got) != str(cell["v"]):
                    problems.append(f'{sheet["name"]}!{addr} value: want {cell["v"]!r} got {xc.value!r}')
            st = styles.get(cell["s"]) if isinstance(cell.get("s"), str) else cell.get("s")
            if st and st.get("n", {}).get("pattern") and xc.number_format != st["n"]["pattern"]:
                problems.append(f'{sheet["name"]}!{addr} numFmt: want {st["n"]["pattern"]} got {xc.number_format}')
    for m in sheet.get("mergeData") or []:
        rng = f'{get_column_letter(m["startColumn"]+1)}{m["startRow"]+1}:{get_column_letter(m["endColumn"]+1)}{m["endRow"]+1}'
        if rng not in [str(x) for x in ws.merged_cells.ranges]:
            problems.append(f'{sheet["name"]}: merge {rng} not applied')

# named ranges
for r in snapshot.get("resources", []):
    if r["name"] == "SHEET_DEFINED_NAME_PLUGIN":
        for d in json.loads(r["data"]).values():
            if d.get("name") and d["name"] not in wb.defined_names:
                problems.append(f'named range "{d["name"]}" missing')

# declared objects
for ch in (objects or {}).get("charts", []):
    ws = wb[ch["sheet"]] if ch["sheet"] in wb.sheetnames else None
    if not ws or not ws._charts:
        problems.append(f'chart "{ch.get("id")}" missing from sheet {ch["sheet"]}')
for pv in (objects or {}).get("pivots", []):
    if pv["sheet"] not in wb.sheetnames:
        problems.append(f'pivot sheet "{pv["sheet"]}" missing')
    else:
        vals = [c.value for col in wb[pv["sheet"]].iter_cols() for c in col]
        if "Grand Total" not in vals:
            problems.append(f'pivot "{pv.get("id")}" has no Grand Total row')

if problems:
    print(f"FAIL — {len(problems)} mismatch(es) over {checked} cells:")
    for p in problems:
        print("  - " + p)
    sys.exit(1)
print(f"PASS — {checked} cells + merges + named ranges + declared objects match the source")
