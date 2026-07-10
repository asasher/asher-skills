#!/usr/bin/env python3
"""conformance.py — measure real round-trip fidelity and gate on it.

For each input .xlsx: import → snapshot(+objects) → re-export → deep-diff original vs round-tripped, and
print a per-file + per-category scorecard (cells compared, cells fully preserved %, mismatches by category).
This turns "fidelity" into a tracked number so every converter fix is measurable, and fails (exit 1) if any
file falls below --min. Point it at real workbooks (e.g. ~/Downloads/Dalia/*.xlsx), not just the sample.

Usage: python3 verify/conformance.py --converter <dir> [--min 0.6] <file.xlsx> [more.xlsx ...]
Stdlib + openpyxl + this skill's xlsx_theme.
"""
import argparse
import os
import subprocess
import sys
import tempfile
import warnings
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

warnings.simplefilter("ignore")

def _theme_resolver(converter_dir):
    sys.path.insert(0, converter_dir)
    from xlsx_theme import read_theme, resolve
    return read_theme, resolve

def _style_tuple(cell, theme, resolve):
    f = cell.font
    fill = cell.fill
    return (
        bool(f and f.bold), bool(f and f.italic),
        (f.size if f else None), (f.name if f else None),
        resolve(f.color, theme) if f else None,                                   # font colour (resolved)
        resolve(fill.fgColor, theme) if fill and fill.patternType == "solid" else None,  # fill (resolved)
        tuple((getattr(cell.border, e).style if getattr(cell.border, e) else None) for e in ("top","bottom","left","right")),
        (cell.alignment.horizontal, cell.alignment.vertical, bool(cell.alignment.wrap_text)) if cell.alignment else None,
        cell.number_format,
    )

def _sheet_visibility(wb):
    return tuple((ws.title, ws.sheet_state) for ws in wb.worksheets)

def _comments(wb):
    return tuple(sorted(
        (ws.title, cell.coordinate, cell.comment.text, cell.comment.author)
        for ws in wb.worksheets for cell in ws._cells.values()
        if getattr(cell, "comment", None) is not None
    ))

def _hyperlinks(wb):
    return tuple(sorted(
        (ws.title, cell.coordinate, cell.hyperlink.target, cell.hyperlink.location,
         cell.hyperlink.display, cell.hyperlink.tooltip)
        for ws in wb.worksheets for cell in ws._cells.values()
        if getattr(cell, "hyperlink", None) is not None
    ))

def _column_span_widths(wb):
    records = []
    for ws in wb.worksheets:
        for key, dim in ws.column_dimensions.items():
            index = column_index_from_string(key)
            records.append((ws.title, dim.min or index, dim.max or index, dim.width))
    return tuple(sorted(records))

def _row_heights(wb):
    return tuple(sorted(
        (ws.title, index, dim.height)
        for ws in wb.worksheets for index, dim in ws.row_dimensions.items()
        if dim.height is not None
    ))

def _exact_structure(original, roundtrip):
    return len(original), len(roundtrip), original == roundtrip

def diff_workbook(orig_path, rt_path, resolve, read_theme, cap=200000):
    o_theme, r_theme = read_theme(orig_path), read_theme(rt_path)
    o_f = load_workbook(orig_path, data_only=False); o_v = load_workbook(orig_path, data_only=True)
    r_f = load_workbook(rt_path, data_only=False); r_v = load_workbook(rt_path, data_only=True)
    cats = {k: 0 for k in ("literal_values","formulas","cached_values","number_formats","fonts","fills","borders","alignment","styles_overall")}
    compared = preserved = 0
    truncated = False
    for name in o_f.sheetnames:
        if name not in r_f.sheetnames:
            continue
        os_, rs = o_f[name], r_f[name]
        ov, rv = (o_v[name] if name in o_v.sheetnames else None), (r_v[name] if name in r_v.sheetnames else None)
        for row in os_.iter_rows():
            for oc in row:
                if oc.value is None:   # value-bearing cells only (comparable to the gpt-5.6 baseline)
                    continue
                if compared >= cap:
                    truncated = True
                    break
                rc = rs.cell(row=oc.row, column=oc.column)
                compared += 1
                ok = True
                # values / formulas
                if isinstance(oc.value, str) and oc.value.startswith("="):
                    if rc.value != oc.value: cats["formulas"] += 1; ok = False
                    ocv = ov.cell(row=oc.row, column=oc.column).value if ov else None
                    rcv = rv.cell(row=oc.row, column=oc.column).value if rv else None
                    if ocv is not None and str(ocv) != str(rcv): cats["cached_values"] += 1  # not fatal to "preserved"
                else:
                    if str(oc.value) != str(rc.value): cats["literal_values"] += 1; ok = False
                # number format
                if oc.number_format != rc.number_format: cats["number_formats"] += 1; ok = False
                # style attributes (resolved colours)
                ot, rt_ = _style_tuple(oc, o_theme, resolve), _style_tuple(rc, r_theme, resolve)
                if ot[4] != rt_[4]: cats["fonts"] += 1; ok = False
                if ot[5] != rt_[5]: cats["fills"] += 1; ok = False
                if ot[6] != rt_[6]: cats["borders"] += 1; ok = False
                if ot[7] != rt_[7]: cats["alignment"] += 1; ok = False
                if ot != rt_: cats["styles_overall"] += 1
                if ok: preserved += 1
            if truncated:
                break
    def _dv(wb):
        return sum(len(ws.data_validations.dataValidation) for ws in wb.worksheets)
    def _cf(wb):
        return sum(len(ws.conditional_formatting._cf_rules) for ws in wb.worksheets)
    def _merges(wb):
        return sum(len(ws.merged_cells.ranges) for ws in wb.worksheets)
    struct = {
        "sheets": (len(o_f.sheetnames), len(r_f.sheetnames), len(o_f.sheetnames) == len(r_f.sheetnames)),
        "names": (len(list(o_f.defined_names)), len(list(r_f.defined_names)),
                  len(list(o_f.defined_names)) == len(list(r_f.defined_names))),
        "merges": (_merges(o_f), _merges(r_f), _merges(o_f) == _merges(r_f)),
        "data_validation": (_dv(o_f), _dv(r_f), _dv(o_f) == _dv(r_f)),
        "cond_format": (_cf(o_f), _cf(r_f), _cf(o_f) == _cf(r_f)),
        "visibility": _exact_structure(_sheet_visibility(o_f), _sheet_visibility(r_f)),
        "comments": _exact_structure(_comments(o_f), _comments(r_f)),
        "hyperlinks": _exact_structure(_hyperlinks(o_f), _hyperlinks(r_f)),
        "column_spans": _exact_structure(_column_span_widths(o_f), _column_span_widths(r_f)),
        "row_heights": _exact_structure(_row_heights(o_f), _row_heights(r_f)),
    }
    return compared, preserved, cats, struct, truncated

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--converter", required=True)
    ap.add_argument("--min", type=float, default=0.0)
    ap.add_argument("files", nargs="+")
    args = ap.parse_args()
    conv = os.path.abspath(args.converter)
    read_theme, resolve = _theme_resolver(conv)
    imp = os.path.join(conv, "xlsx-to-snapshot.py")
    exp = os.path.join(conv, "snapshot-to-xlsx.py")

    overall_ok = True
    print("note: 'preserved' = value+formula+numfmt+resolved-style match. cached = formula cells whose")
    print("      cached result was dropped (backlog #3). colour compare is pipeline-self-consistent, not a")
    print("      pixel ground-truth — it catches 'colour missing', not a resolver tint error.\n")
    print(f"{'workbook':<30}{'cells':>8}{'preserved':>10}{'%':>7}{'cached-lost':>12}  top style losses")
    print("-" * 100)
    for path in args.files:
        base = os.path.splitext(os.path.basename(path))[0]
        with tempfile.TemporaryDirectory() as td:
            snap = os.path.join(td, "s.json"); obj = os.path.join(td, "o.json"); rt = os.path.join(td, "rt.xlsx")
            ri = subprocess.run([sys.executable, imp, path, snap, obj], capture_output=True, text=True)
            if not os.path.exists(snap):
                print(f"{base[:33]:<34}{'IMPORT FAILED':>28}  {ri.stderr.strip().splitlines()[-1][:40] if ri.stderr else ''}")
                overall_ok = False; continue
            re_ = subprocess.run([sys.executable, exp, snap, rt, obj], capture_output=True, text=True)
            if not os.path.exists(rt):
                print(f"{base[:33]:<34}{'EXPORT FAILED':>28}  {re_.stderr.strip().splitlines()[-1][:40] if re_.stderr else ''}")
                overall_ok = False; continue
            compared, preserved, cats, struct, trunc = diff_workbook(path, rt, resolve, read_theme)
            pct = preserved / compared if compared else 0.0
            style_cats = ("fonts", "fills", "borders", "alignment", "number_formats", "literal_values", "formulas")
            top = sorted(((cats[k], k) for k in style_cats if cats[k]), reverse=True)[:3]
            losses = ", ".join(f"{k}={v}" for v, k in top) or "none"
            struct_ok = all(match for _, _, match in struct.values())
            flag = ("" if pct >= args.min else "  ⛔<min") + ("" if struct_ok else "  ⛔structure")
            note = " (capped)" if trunc else ""
            print(f"{base[:29]:<30}{compared:>8}{preserved:>10}{pct:>6.1%}{cats['cached_values']:>12}  {losses}{note}{flag}")
            st = "  structural (orig→rt): " + " · ".join(
                f"{k} {a}→{b}{'✓' if match else '✗'}" for k, (a, b, match) in struct.items())
            print(st)
            if pct < args.min or not struct_ok:
                overall_ok = False
    print("-" * 96)
    print(f"gate: {'PASS' if overall_ok else 'FAIL'} (min={args.min:.0%})")
    sys.exit(0 if overall_ok else 1)

if __name__ == "__main__":
    main()
