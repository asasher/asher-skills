#!/usr/bin/env python3
"""snapshot-to-xlsx.py — compile a Univer snapshot (+ declared objects) to .xlsx via openpyxl.

Source of truth: workbook.snapshot.json (Univer-native cells) + objects.json (declared charts/pivots
that Univer OSS can't natively hold). openpyxl gives us native charts, all conditional-formatting types,
validation, images, and named ranges for free. Pivots are flattened to a static computed table here; a
higher-fidelity live pivot is materialized by the optional LibreOffice pass (materialize-pivots.py).

Committed feature set: values, formulas, styles, merges, exact sizing/spans, sheet visibility, comments,
hyperlinks, images, freeze, named ranges, validation, conditional formatting, native charts, flattened pivots.

Usage: python3 snapshot-to-xlsx.py <snapshot.json> <out.xlsx> [objects.json]
Stdlib + openpyxl only.
"""
import base64
import datetime as dt
import io
import json
import os
import re
import sys
import tempfile
import zipfile
from xml.sax.saxutils import escape
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor, AnchorMarker, OneCellAnchor, TwoCellAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.datetime import to_excel
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.formatting.rule import Rule, FormatObject, ColorScale, DataBar, IconSet
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.cell.cell import ERROR_CODES

# ---- normalizers ---------------------------------------------------------------------------
def _parse_iso(s):
    """ISO date/time string -> datetime/date/time (falls back to the raw string if unparseable)."""
    for parser in (dt.datetime.fromisoformat, dt.date.fromisoformat, dt.time.fromisoformat):
        try:
            return parser(s)
        except (ValueError, TypeError):
            continue
    return s

def to_argb(color):
    if not color:
        return None
    raw = color.get("rgb") if isinstance(color, dict) else color
    if not raw:
        return None
    m = re.match(r"rgba?\(([^)]+)\)", raw, re.I)
    if m:
        r, g, b = (int(x.strip()) for x in m.group(1).split(",")[:3])
        hexs = f"{r:02X}{g:02X}{b:02X}"
    else:
        hexs = raw.lstrip("#")
    if len(hexs) == 3:
        hexs = "".join(c * 2 for c in hexs)
    return ("FF" + hexs).upper()

BORDER_STYLE = {1: "thin", 2: "hair", 3: "dotted", 4: "dashed", 5: "dashDot", 6: "dashDotDot",
                7: "double", 8: "medium", 9: "mediumDashed", 10: "mediumDashDot", 11: "mediumDashDotDot",
                12: "slantDashDot", 13: "thick"}
H_ALIGN = {1: "left", 2: "center", 3: "right"}
V_ALIGN = {1: "top", 2: "center", 3: "bottom"}

def apply_style(cell, s):
    if not s:
        return
    font_kw = {}
    if s.get("bl"): font_kw["bold"] = True
    if s.get("it"): font_kw["italic"] = True
    if s.get("ul"): font_kw["underline"] = "single"
    if s.get("st"): font_kw["strike"] = True
    if s.get("fs"): font_kw["size"] = s["fs"]
    if s.get("ff"): font_kw["name"] = s["ff"]
    if s.get("cl"): font_kw["color"] = to_argb(s["cl"])
    if font_kw: cell.font = Font(**font_kw)
    if s.get("bg"): cell.fill = PatternFill(fill_type="solid", fgColor=to_argb(s["bg"]))
    if s.get("bd"):
        sides = {}
        for uk, ek in (("t", "top"), ("b", "bottom"), ("l", "left"), ("r", "right")):
            edge = s["bd"].get(uk)
            if edge and edge.get("s"):
                sides[ek] = Side(style=BORDER_STYLE.get(edge["s"], "thin"),
                                 color=to_argb(edge.get("cl")) or "FF000000")
        if sides: cell.border = Border(**sides)
    align_kw = {}
    if s.get("ht") in H_ALIGN: align_kw["horizontal"] = H_ALIGN[s["ht"]]
    if s.get("vt") in V_ALIGN: align_kw["vertical"] = V_ALIGN[s["vt"]]
    if s.get("tb") == 3: align_kw["wrap_text"] = True
    if align_kw: cell.alignment = Alignment(**align_kw)
    n = s.get("n")
    if n and n.get("pattern"): cell.number_format = n["pattern"]

def _font_from_style(s):
    font_kw = {}
    if s.get("bl"): font_kw["bold"] = True
    if s.get("it"): font_kw["italic"] = True
    if s.get("ul"): font_kw["underline"] = "single"
    if s.get("st"): font_kw["strike"] = True
    if s.get("fs") is not None: font_kw["size"] = s["fs"]
    if s.get("ff"): font_kw["name"] = s["ff"]
    if s.get("cl"): font_kw["color"] = to_argb(s["cl"])
    return Font(**font_kw) if font_kw else None

def _dxf_from_style(s):
    if not s:
        return None
    fill = PatternFill(fill_type="solid", fgColor=to_argb(s["bg"])) if s.get("bg") else None
    font = _font_from_style(s)
    return DifferentialStyle(font=font, fill=fill) if font or fill else None

# ---- A1 helpers ----------------------------------------------------------------------------
def parse_ref(ref):
    """'Sales!B3:B6' or 'B3:B6' or 'F3' -> (sheet|None, min_r, min_c, max_r, max_c) 1-indexed."""
    sheet = None
    if "!" in ref:
        sheet, ref = ref.split("!", 1)
        sheet = sheet.strip("'")
    parts = ref.replace("$", "").split(":")
    def rc(a1):
        m = re.match(r"([A-Za-z]+)(\d+)", a1)
        return int(m.group(2)), column_index_from_string(m.group(1))
    r1, c1 = rc(parts[0])
    r2, c2 = rc(parts[1]) if len(parts) > 1 else (r1, c1)
    return sheet, min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2)

def _covered(column, spans):
    return any(start <= column <= end for start, end in spans)

def _apply_dimensions(ws, sheet):
    """Prefer exact native measures/spans; retain legacy pixel fields as a back-compat fallback."""
    spans = []
    for data in sheet.get("columnDimensions") or []:
        start = int(data["startColumn"])
        end = int(data.get("endColumn", start))
        if end < start:
            raise ValueError(f"invalid column span {start}:{end} on {sheet.get('name')!r}")
        dim = ws.column_dimensions[get_column_letter(start + 1)]
        dim.min, dim.max = start + 1, end + 1
        if data.get("width") is not None:
            dim.width = data["width"]
        if data.get("hidden"):
            dim.hidden = True
        if data.get("outlineLevel") is not None:
            dim.outlineLevel = data["outlineLevel"]
        if data.get("collapsed"):
            dim.collapsed = True
        if data.get("bestFit"):
            dim.bestFit = True
        spans.append((start, end))

    for c, data in (sheet.get("columnData") or {}).items():
        column = int(c)
        dim = ws.column_dimensions[get_column_letter(column + 1)]
        if not _covered(column, spans):
            if data.get("width") is not None:
                dim.width = data["width"]
            elif data.get("w") is not None:
                dim.width = data["w"] / 7
        if data.get("hd"):
            dim.hidden = True

    for r, data in (sheet.get("rowData") or {}).items():
        dim = ws.row_dimensions[int(r) + 1]
        if data.get("height") is not None:
            dim.height = data["height"]
        elif data.get("h") is not None:
            dim.height = data["h"] * 0.75
        if data.get("hd"):
            dim.hidden = True
        if data.get("outlineLevel") is not None:
            dim.outlineLevel = data["outlineLevel"]
        if data.get("collapsed"):
            dim.collapsed = True

def _marker(data):
    return AnchorMarker(col=int(data["column"]), colOff=int(data.get("columnOffset", 0)),
                        row=int(data["row"]), rowOff=int(data.get("rowOffset", 0)))

def _drawing_anchor(data):
    data = data or {"type": "cell", "coordinate": "A1"}
    kind = data.get("type", "cell")
    if kind == "cell":
        return data.get("coordinate", "A1")
    if kind == "oneCell":
        extent = data["extent"]
        return OneCellAnchor(_from=_marker(data["from"]),
                             ext=XDRPositiveSize2D(cx=extent["width"], cy=extent["height"]))
    if kind == "twoCell":
        return TwoCellAnchor(editAs=data.get("editAs"), _from=_marker(data["from"]),
                             to=_marker(data["to"]))
    if kind == "absolute":
        position, extent = data["position"], data["extent"]
        return AbsoluteAnchor(pos=XDRPoint2D(x=position["x"], y=position["y"]),
                              ext=XDRPositiveSize2D(cx=extent["width"], cy=extent["height"]))
    raise ValueError(f"unsupported image anchor type {kind!r}")

def _apply_images(ws, resources):
    for index, resource in enumerate(resources or []):
        try:
            raw = base64.b64decode(resource["data"], validate=True)
            image = Image(io.BytesIO(raw))
            if resource.get("width") is not None:
                image.width = resource["width"]
            if resource.get("height") is not None:
                image.height = resource["height"]
            ws.add_image(image, _drawing_anchor(resource.get("anchor")))
        except Exception as exc:
            raise ValueError(f"{ws.title} image {index + 1} could not be restored: {exc}") from exc

# ---- snapshot -> workbook ------------------------------------------------------------------
def build(snapshot, objects):
    wb = Workbook()
    wb.remove(wb.active)
    styles = snapshot.get("styles", {})
    resolve = lambda s: styles.get(s) if isinstance(s, str) else s
    order = snapshot.get("sheetOrder") or list(snapshot.get("sheets", {}))
    ws_by_name = {}
    ws_by_id = {}

    for sid in order:
        sheet = snapshot["sheets"].get(sid)
        if not sheet:
            continue
        ws = wb.create_sheet(title=sheet["name"])
        ws_by_name[sheet["name"]] = ws
        ws_by_id[sid] = ws
        state = sheet.get("sheetState")
        if state in {"visible", "hidden", "veryHidden"}:
            ws.sheet_state = state

        _apply_dimensions(ws, sheet)

        for r, row in (sheet.get("cellData") or {}).items():
            for c, cell in row.items():
                if cell is None:
                    continue
                xc = ws.cell(row=int(r) + 1, column=int(c) + 1)
                if cell.get("f"):
                    xc.value = cell["f"] if cell["f"].startswith("=") else "=" + cell["f"]
                elif cell.get("t") == "d" and isinstance(cell.get("v"), str):
                    xc.value = _parse_iso(cell["v"])   # re-materialize date/time (paired with import's "d" marker)
                elif cell.get("t") == "s" and isinstance(cell.get("v"), str):
                    xc.value = cell["v"]
                    xc.data_type = "s"   # force literal: a string starting with '=' would else become a formula
                elif cell.get("v") is not None:
                    xc.value = cell["v"]
                apply_style(xc, resolve(cell.get("s")))
                comment = cell.get("comment")
                if isinstance(comment, dict) and comment.get("text") is not None:
                    xc.comment = Comment(comment["text"], comment.get("author") or "")
                link = cell.get("hyperlink")
                if isinstance(link, dict) and (link.get("target") or link.get("location")):
                    target = link.get("target")
                    location = link.get("location")
                    if location is not None and target == location:
                        target = None
                    xc.hyperlink = Hyperlink(ref=xc.coordinate, target=target, location=location,
                                             display=link.get("display"), tooltip=link.get("tooltip"))

        for m in sheet.get("mergeData") or []:
            ws.merge_cells(start_row=m["startRow"] + 1, start_column=m["startColumn"] + 1,
                           end_row=m["endRow"] + 1, end_column=m["endColumn"] + 1)

        fr = sheet.get("freeze")
        if fr and (fr.get("xSplit") or fr.get("ySplit")):
            ws.freeze_panes = ws.cell(row=(fr.get("ySplit") or 0) + 1, column=(fr.get("xSplit") or 0) + 1)

        _apply_images(ws, sheet.get("images"))

    _apply_named_ranges(snapshot, wb)
    _apply_validation(snapshot, ws_by_name, wb)
    _apply_cond_format(snapshot, ws_by_name)
    _apply_charts(objects, ws_by_name, ws_by_id)
    _apply_pivots(objects, snapshot, wb, ws_by_name, ws_by_id)
    return wb

def _resource(snapshot, name):
    for r in snapshot.get("resources", []):
        if r.get("name") == name:
            data = r.get("data")
            return json.loads(data) if isinstance(data, str) else data
    return None

def _sheet_name_by_id(snapshot, sid):
    return snapshot["sheets"][sid]["name"]

def _validation_formula(rule_type, value, defined_names):
    if value is None or rule_type != "list":
        return value
    formula = str(value)
    if len(formula) >= 2 and formula.startswith('"') and formula.endswith('"'):
        return formula
    is_reference = "!" in formula or formula.startswith("=") or formula in defined_names
    if "," in formula and not is_reference:
        return '"%s"' % formula.replace('"', '""')
    return formula

def _apply_validation(snapshot, ws_by_name, wb):
    dv = _resource(snapshot, "SHEET_DATA_VALIDATION_PLUGIN") or {}
    supported = {"list", "whole", "decimal", "date", "textLength", "custom"}
    defined_names = set(wb.defined_names)
    for ws in wb.worksheets:
        defined_names.update(ws.defined_names)
    for sid, rules in dv.items():
        ws = ws_by_name.get(_sheet_name_by_id(snapshot, sid))
        if not ws:
            continue
        for rule in rules:
            rule_type = rule.get("type")
            if rule_type not in supported:
                continue
            formula1 = _validation_formula(rule_type, rule.get("formula1"), defined_names)
            d = DataValidation(type=rule_type, operator=rule.get("operator"), formula1=formula1,
                               formula2=rule.get("formula2"), allow_blank=rule.get("allowBlank", True))
            ws.add_data_validation(d)
            for rg in rule.get("ranges", []):
                d.add("%s%d:%s%d" % (get_column_letter(rg["startColumn"] + 1), rg["startRow"] + 1,
                                     get_column_letter(rg["endColumn"] + 1), rg["endRow"] + 1))

def _format_object(point, default_type, default_value=None):
    point = point or {}
    return FormatObject(type=point.get("type", default_type), val=point.get("value", default_value),
                        gte=point.get("gte"))

def _rule_kwargs(rule):
    return {"priority": rule.get("priority", 0), "stopIfTrue": rule.get("stopIfTrue"),
            "dxf": _dxf_from_style(rule.get("style"))}

def _apply_cond_format(snapshot, ws_by_name):
    cf = _resource(snapshot, "SHEET_CONDITIONAL_FORMATTING_PLUGIN") or {}
    for sid, items in cf.items():
        ws = ws_by_name.get(_sheet_name_by_id(snapshot, sid))
        if not ws:
            continue
        for item in items:
            rule = item.get("rule", {})
            ref = " ".join("%s%d:%s%d" % (get_column_letter(rg["startColumn"] + 1), rg["startRow"] + 1,
                                          get_column_letter(rg["endColumn"] + 1), rg["endRow"] + 1)
                           for rg in item.get("ranges", []))
            t = rule.get("type")
            if t == "highlightCell":
                formulas = []
                if rule.get("value") is not None:
                    formulas.append(str(rule["value"]))
                if rule.get("value2") is not None:
                    formulas.append(str(rule["value2"]))
                ws.conditional_formatting.add(ref, Rule(
                    type="cellIs", operator=rule.get("operator", "greaterThan"), formula=formulas,
                    **_rule_kwargs(rule)))
            elif t == "colorScale":
                if rule.get("config"):
                    config = sorted(rule["config"], key=lambda point: point.get("index", 0))
                    points = [(point.get("value") or {},
                               "min" if i == 0 else "max" if i == len(config) - 1 else "percentile")
                              for i, point in enumerate(config)]
                    colors = [Color(rgb=to_argb(point.get("color"))) for point in config]
                else:
                    points = [(rule.get("min"), "min")]
                    if rule.get("mid"):
                        points.append((rule["mid"], "percentile"))
                    points.append((rule.get("max"), "max"))
                    colors = [Color(rgb=to_argb((point or {}).get("color"))) for point, _ in points]
                cfvo = [_format_object(point, default, 50 if default == "percentile" else None)
                        for point, default in points]
                ws.conditional_formatting.add(ref, Rule(
                    type="colorScale", colorScale=ColorScale(cfvo=cfvo, color=colors),
                    **_rule_kwargs(rule)))
            elif t == "dataBar":
                bar = DataBar(cfvo=[_format_object(rule.get("min"), "min"),
                                    _format_object(rule.get("max"), "max")],
                              color=Color(rgb=to_argb(rule.get("color") or "#638EC6")),
                              showValue=rule.get("showValue"), minLength=rule.get("minLength"),
                              maxLength=rule.get("maxLength"))
                ws.conditional_formatting.add(ref, Rule(type="dataBar", dataBar=bar,
                                                        **_rule_kwargs(rule)))
            elif t == "iconSet":
                thresholds = rule.get("thresholds")
                if not thresholds:
                    icon_count = int(str(rule.get("iconStyle", "3TrafficLights1"))[0])
                    thresholds = [{"type": "percent", "value": round(i * 100 / icon_count)}
                                  for i in range(icon_count)]
                icons = IconSet(iconSet=rule.get("iconStyle", "3TrafficLights1"),
                                cfvo=[_format_object(point, "percent") for point in thresholds],
                                showValue=rule.get("showValue"), percent=rule.get("percent"),
                                reverse=rule.get("reverse"))
                ws.conditional_formatting.add(ref, Rule(type="iconSet", iconSet=icons,
                                                        **_rule_kwargs(rule)))

def _apply_named_ranges(snapshot, wb):
    dn = _resource(snapshot, "SHEET_DEFINED_NAME_PLUGIN") or {}
    for d in dn.values():
        if d.get("name") and d.get("formulaOrRefString"):
            ref = d["formulaOrRefString"].lstrip("=")
            local_sheet_id = d.get("localSheetId")
            defn = DefinedName(name=d["name"], attr_text=ref, localSheetId=local_sheet_id)
            try:
                if local_sheet_id is None:
                    wb.defined_names.add(defn)
                else:
                    wb.worksheets[int(local_sheet_id)].defined_names.add(defn)
            except Exception:
                if local_sheet_id is None:
                    wb.defined_names[d["name"]] = defn
                else:
                    raise

_CHART = {"bar": BarChart, "line": LineChart, "pie": PieChart}

def _object_sheet(obj, ws_by_name, ws_by_id):
    if obj.get("sheetId") is not None:
        return ws_by_id.get(obj["sheetId"])
    return ws_by_name.get(obj.get("sheet"))

def _apply_charts(objects, ws_by_name, ws_by_id):
    for ch in (objects or {}).get("charts", []):
        ws = _object_sheet(ch, ws_by_name, ws_by_id)
        if not ws:
            raise ValueError(f"chart {ch.get('id', '')}: target sheet could not be resolved")
        chart = _CHART.get(ch.get("type", "bar"), BarChart)()
        if ch.get("title"): chart.title = ch["title"]
        for vref in ch.get("values", []):
            source_sheet, r1, c1, r2, c2 = parse_ref(vref)
            source_ws = ws_by_name.get(source_sheet) if source_sheet else ws
            if not source_ws:
                continue
            data = Reference(source_ws, min_col=c1, min_row=r1, max_col=c2, max_row=r2)
            chart.add_data(data, titles_from_data=False)
        if ch.get("categories"):
            source_sheet, r1, c1, r2, c2 = parse_ref(ch["categories"])
            source_ws = ws_by_name.get(source_sheet) if source_sheet else ws
            if source_ws:
                chart.set_categories(Reference(source_ws, min_col=c1, min_row=r1, max_col=c2, max_row=r2))
        for i, title in enumerate(ch.get("seriesTitles", [])):
            if i < len(chart.series):
                chart.series[i].tx = SeriesLabel(v=title)
        ws.add_chart(chart, ch.get("anchor", "H2"))

def _cell_value(snapshot, sheet_name, row1, col1):
    for sid, sheet in snapshot["sheets"].items():
        if sheet["name"] != sheet_name:
            continue
        cell = (sheet.get("cellData") or {}).get(str(row1 - 1), {}).get(str(col1 - 1))
        if cell:
            return cell.get("v")
    return None

_AGGREGATIONS = {"sum", "count", "avg", "min", "max"}

def _new_aggregate(values):
    return [{"sum": 0, "count": 0, "min": None, "max": None} for _ in values]

def _update_aggregate(states, row, idx, values):
    for i, spec in enumerate(values):
        value = row[idx[spec["field"]]]
        agg = spec.get("agg", "sum")
        if agg == "count":
            if value is not None:
                states[i]["count"] += 1
            continue
        if not isinstance(value, (int, float)) or isinstance(value, bool):
            continue
        state = states[i]
        state["sum"] += value
        state["count"] += 1
        state["min"] = value if state["min"] is None else min(state["min"], value)
        state["max"] = value if state["max"] is None else max(state["max"], value)

def _aggregate_result(state, agg):
    if agg == "sum":
        return state["sum"]
    if agg == "count":
        return state["count"]
    if agg == "avg":
        return state["sum"] / state["count"] if state["count"] else None
    return state[agg]

def _aggregate_label(spec, qualify=False):
    agg = spec.get("agg", "sum")
    return f"{agg} of {spec['field']}" if qualify else spec["field"]

def _apply_pivots(objects, snapshot, wb, ws_by_name, ws_by_id):
    """Flatten declared row/column pivots with sum/count/avg/min/max aggregation."""
    for pv in (objects or {}).get("pivots", []):
        sheet_name, r1, c1, r2, c2 = parse_ref(pv["source"])
        headers = [_cell_value(snapshot, sheet_name, r1, c) for c in range(c1, c2 + 1)]
        idx = {h: i for i, h in enumerate(headers)}
        rows = []
        for rr in range(r1 + 1, r2 + 1):
            rows.append([_cell_value(snapshot, sheet_name, rr, c) for c in range(c1, c2 + 1)])
        row_fields = pv.get("rows", [])
        col_fields = pv.get("cols", [])
        values = pv.get("values", [])
        for field in row_fields + col_fields + [v["field"] for v in values]:
            if field not in idx:
                raise ValueError(f"pivot {pv.get('id', '')}: field {field!r} not found in source headers")
        for spec in values:
            if spec.get("agg", "sum") not in _AGGREGATIONS:
                raise ValueError(f"pivot {pv.get('id', '')}: unsupported aggregation {spec.get('agg')!r}")

        groups = {}
        row_totals = {}
        col_totals = {}
        grand_total = _new_aggregate(values)
        row_keys, col_keys = set(), set()
        for row in rows:
            row_key = tuple(row[idx[f]] for f in row_fields)
            col_key = tuple(row[idx[f]] for f in col_fields)
            row_keys.add(row_key)
            col_keys.add(col_key)
            for aggregate in (groups.setdefault((row_key, col_key), _new_aggregate(values)),
                              row_totals.setdefault(row_key, _new_aggregate(values)),
                              col_totals.setdefault(col_key, _new_aggregate(values)), grand_total):
                _update_aggregate(aggregate, row, idx, values)

        out = _object_sheet(pv, ws_by_name, ws_by_id)
        if out is None:
            if pv.get("sheetId") is not None:
                raise ValueError(f"pivot {pv.get('id', '')}: target sheetId could not be resolved")
            out = wb.create_sheet(title=pv["sheet"])
            ws_by_name[pv["sheet"]] = out
        anchor_r, anchor_c = parse_ref(pv.get("anchor", "A1"))[1:3]
        sorted_rows = sorted(row_keys, key=lambda k: tuple(str(x) for x in k))
        sorted_cols = sorted(col_keys, key=lambda k: tuple(str(x) for x in k))
        qualify_values = len(values) != len({v["field"] for v in values})
        headers_out = list(row_fields)
        if col_fields:
            for col_key in sorted_cols:
                prefix = " / ".join(str(x) for x in col_key)
                headers_out.extend(f"{prefix} | {_aggregate_label(v, qualify_values)}" for v in values)
            headers_out.extend(f"Grand Total | {_aggregate_label(v, qualify_values)}" for v in values)
        else:
            headers_out.extend(_aggregate_label(v, qualify_values) for v in values)
        for j, f in enumerate(headers_out):
            hc = out.cell(row=anchor_r, column=anchor_c + j, value=f)
            hc.font = Font(bold=True)

        rownum = anchor_r + 1
        for row_key in sorted_rows:
            for j, kv in enumerate(row_key):
                out.cell(row=rownum, column=anchor_c + j, value=kv)
            offset = len(row_fields)
            if col_fields:
                for col_key in sorted_cols:
                    states = groups.get((row_key, col_key), _new_aggregate(values))
                    for j, spec in enumerate(values):
                        out.cell(row=rownum, column=anchor_c + offset,
                                 value=_aggregate_result(states[j], spec.get("agg", "sum")))
                        offset += 1
                states = row_totals[row_key]
            else:
                states = groups.get((row_key, ()), _new_aggregate(values))
            for j, spec in enumerate(values):
                out.cell(row=rownum, column=anchor_c + offset + j,
                         value=_aggregate_result(states[j], spec.get("agg", "sum")))
            rownum += 1
        gt = out.cell(row=rownum, column=anchor_c, value="Grand Total")
        gt.font = Font(bold=True)
        offset = len(row_fields)
        if col_fields:
            for col_key in sorted_cols:
                states = col_totals[col_key]
                for j, spec in enumerate(values):
                    tc = out.cell(row=rownum, column=anchor_c + offset,
                                  value=_aggregate_result(states[j], spec.get("agg", "sum")))
                    tc.font = Font(bold=True)
                    offset += 1
        for j, spec in enumerate(values):
            tc = out.cell(row=rownum, column=anchor_c + offset + j,
                          value=_aggregate_result(grand_total[j], spec.get("agg", "sum")))
            tc.font = Font(bold=True)

def _cached_xml_value(cell, epoch):
    value = cell.get("v")
    if cell.get("t") == "d" and isinstance(value, str):
        value = _parse_iso(value)
        if isinstance(value, (dt.datetime, dt.date, dt.time)):
            return str(to_excel(value, epoch)), None
    if isinstance(value, bool):
        return ("1" if value else "0"), "b"
    if isinstance(value, (int, float)):
        return str(value), None
    if cell.get("t") == "e" or value in ERROR_CODES:
        return str(value), "e"
    return str(value), "str"

def _set_cached_value(xml, address, value, cell_type):
    address_b = re.escape(address.encode("ascii"))
    cell_re = re.compile(rb'(<c\b(?=[^>]*\br="' + address_b + rb'")[^>]*>)(.*?)(</c>)', re.S)

    def replace(match):
        tag, body, close = match.groups()
        formula = re.search(rb'<f(?:\s[^>]*)?>.*?</f>', body, re.S)
        if not formula:
            return match.group(0)
        tag = re.sub(rb'\s+t="[^"]*"', b'', tag)
        if cell_type:
            tag = tag[:-1] + b' t="' + cell_type.encode("ascii") + b'">'
        remainder = body[formula.end():]
        remainder = re.sub(rb'^\s*<v(?:\s[^>]*)?>.*?</v>', b'', remainder, count=1, flags=re.S)
        cached = b'<v>' + escape(value).encode("utf-8") + b'</v>'
        return tag + body[:formula.end()] + cached + remainder + close

    return cell_re.sub(replace, xml, count=1)

def _inject_formula_caches(path, snapshot, epoch):
    """Atomically replace openpyxl's empty formula <v> nodes with snapshot cached values."""
    caches = {}
    sheet_number = 0
    order = snapshot.get("sheetOrder") or list(snapshot.get("sheets", {}))
    for sid in order:
        sheet = snapshot.get("sheets", {}).get(sid)
        if not sheet:
            continue
        sheet_number += 1
        values = {}
        for row, cells in (sheet.get("cellData") or {}).items():
            for col, cell in cells.items():
                if cell and cell.get("f") and "v" in cell and cell.get("v") is not None:
                    address = f"{get_column_letter(int(col) + 1)}{int(row) + 1}"
                    values[address] = _cached_xml_value(cell, epoch)
        if values:
            caches[f"xl/worksheets/sheet{sheet_number}.xml"] = values
    if not caches:
        return

    directory = os.path.dirname(os.path.abspath(path))
    fd, tmp = tempfile.mkstemp(prefix=".xlsx-cache-", suffix=".xlsx", dir=directory)
    os.close(fd)
    try:
        with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(tmp, "w") as target:
            for info in source.infolist():
                data = source.read(info.filename)
                for address, (value, cell_type) in caches.get(info.filename, {}).items():
                    data = _set_cached_value(data, address, value, cell_type)
                target.writestr(info, data)
        os.chmod(tmp, os.stat(path).st_mode)
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)

def _closure_gate(objects, snapshot):
    """Closure gate: refuse to compile invalid declared objects (schema/reference errors), but only WARN on
    undeclared import gaps (an imported workbook with charts/pivots you chose not to re-declare)."""
    try:
        from validate_objects import validate_objects, validate_against_snapshot
    except ImportError:
        return  # validator optional
    errs = validate_objects(objects) + validate_against_snapshot(objects, snapshot)
    gaps = [e for e in errs if e.startswith("$._import_note")]
    hard = [e for e in errs if not e.startswith("$._import_note")]
    for g in gaps:
        print("  ⚠ closure gap — " + g, file=sys.stderr)
    if hard:
        print("compile refused — objects.json failed the closure gate:", file=sys.stderr)
        for e in hard:
            print("  - " + e, file=sys.stderr)
        sys.exit(1)

def main():
    if len(sys.argv) < 3:
        print("usage: snapshot-to-xlsx.py <snapshot.json> <out.xlsx> [objects.json]", file=sys.stderr)
        sys.exit(2)
    snapshot = json.load(open(sys.argv[1]))
    out = sys.argv[2]
    objects = json.load(open(sys.argv[3])) if len(sys.argv) > 3 else {}
    if objects:
        _closure_gate(objects, snapshot)
    wb = build(snapshot, objects)
    wb.save(out)
    _inject_formula_caches(out, snapshot, wb.epoch)
    print(f"wrote {out}")

if __name__ == "__main__":
    main()
