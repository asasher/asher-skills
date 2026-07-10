#!/usr/bin/env python3
"""xlsx-to-snapshot.py — import an existing .xlsx into a Univer snapshot (+ objects.json), via openpyxl.

Recovers cells, formulas (including cached results), styles, merges, exact sizing/spans, comments,
hyperlinks, images, sheet visibility, freeze, named ranges, data validation, and conditional formatting.
Charts and pivot tables that openpyxl can't reconstruct
are DETECTED by scanning the .xlsx parts and reported honestly (never silently dropped) — so you know to
re-declare them in objects.json. This closes the "import has X but the browser can't show it" gap by
making X visible as a reported gap rather than a silent loss.

Usage: python3 xlsx-to-snapshot.py <in.xlsx> <out.snapshot.json> [out.objects.json]
Stdlib + openpyxl only.
"""
import base64
import datetime as dt
import json
import os
import re
import sys
import zipfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from xlsx_theme import read_theme, resolve
try:
    from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
except ImportError:  # older openpyxl
    ArrayFormula = DataTableFormula = ()

def _atomic_write_json(path, obj):
    """Write JSON to a temp file then rename, so a crash can't leave a truncated, plausible-looking file.
    default=str is a last-resort safety net: any openpyxl scalar we didn't map explicitly degrades to its
    string form instead of aborting the entire import."""
    tmp = f"{path}.tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, indent=2, default=str)
    os.replace(tmp, path)

BORDER = {"thin": 1, "hair": 2, "dotted": 3, "dashed": 4, "dashDot": 5, "dashDotDot": 6, "double": 7,
          "medium": 8, "mediumDashed": 9, "mediumDashDot": 10, "mediumDashDotDot": 11, "slantDashDot": 12, "thick": 13}
H = {"left": 1, "center": 2, "right": 3}
V = {"top": 1, "center": 2, "bottom": 3}

def style_from(cell, theme_idx):
    """Build a Univer IStyleData from an openpyxl cell. Colours are resolved through the theme/indexed
    palette (theme_idx) so theme colours survive as concrete RGB rather than being dropped."""
    s = {}
    f = cell.font
    if f:
        if f.bold: s["bl"] = 1
        if f.italic: s["it"] = 1
        if f.underline: s["ul"] = {"s": 1}
        if f.strike: s["st"] = {"s": 1}
        if f.size: s["fs"] = f.size            # always keep name+size: omitting them made the export carry
        if f.name: s["ff"] = f.name            # no explicit font, so LibreOffice fell back to a serif face
        c = resolve(f.color, theme_idx)   # keep explicit colours incl. black — dropping black was a bug
        if c: s["cl"] = {"rgb": c}
    fill = cell.fill
    if fill and fill.patternType == "solid":
        c = resolve(fill.fgColor, theme_idx)
        if c: s["bg"] = {"rgb": c}
    bd = {}
    for ek, uk in (("top", "t"), ("bottom", "b"), ("left", "l"), ("right", "r")):
        side = getattr(cell.border, ek, None)
        if side and side.style:
            bd[uk] = {"s": BORDER.get(side.style, 1), "cl": {"rgb": resolve(side.color, theme_idx) or "#000000"}}
    if bd: s["bd"] = bd
    a = cell.alignment
    if a:
        if a.horizontal in H: s["ht"] = H[a.horizontal]
        if a.vertical in V: s["vt"] = V[a.vertical]
        if a.wrap_text: s["tb"] = 3
    if cell.number_format and cell.number_format != "General":
        s["n"] = {"pattern": cell.number_format}
    return s or None

def _dxf_style(dxf, theme_idx):
    """Map the fill and full differential font used by a conditional-formatting rule."""
    if not dxf:
        return None
    s = {}
    f = dxf.font
    if f:
        if f.bold: s["bl"] = 1
        if f.italic: s["it"] = 1
        if f.underline: s["ul"] = {"s": 1}
        if f.strike: s["st"] = {"s": 1}
        if f.size is not None: s["fs"] = f.size
        if f.name: s["ff"] = f.name
        c = resolve(f.color, theme_idx)
        if c: s["cl"] = {"rgb": c}
    fill = dxf.fill
    if fill and fill.patternType == "solid":
        c = resolve(fill.fgColor, theme_idx)
        if c: s["bg"] = {"rgb": c}
    return s or None

def _range_to_iranges(sqref):
    out = []
    for part in str(sqref).split():
        m = re.match(r"([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?", part)
        if not m:
            continue
        c1 = column_index_from_string(m.group(1)) - 1
        r1 = int(m.group(2)) - 1
        c2 = column_index_from_string(m.group(3)) - 1 if m.group(3) else c1
        r2 = int(m.group(4)) - 1 if m.group(4) else r1
        out.append({"startRow": r1, "endRow": r2, "startColumn": c1, "endColumn": c2})
    return out

def _cfvo(obj):
    out = {"type": obj.type}
    if obj.val is not None:
        out["value"] = obj.val
    if obj.gte is not None:
        out["gte"] = obj.gte
    return out

def _with_cf_metadata(out, rule, theme_idx):
    out["priority"] = rule.priority
    out["stopIfTrue"] = rule.stopIfTrue
    style = _dxf_style(rule.dxf, theme_idx)
    if style:
        out["style"] = style
    return out

def read_cond_format(ws, theme_idx):
    """Reverse-map cellIs, colorScale, dataBar, and iconSet rules without flattening thresholds."""
    items = []
    for cf in ws.conditional_formatting:
        ranges = _range_to_iranges(cf.sqref)
        for rule in cf.rules:
            if rule.type == "colorScale" and rule.colorScale:
                points = []
                for i, obj in enumerate(rule.colorScale.cfvo):
                    point = _cfvo(obj)
                    if i < len(rule.colorScale.color):
                        point["color"] = resolve(rule.colorScale.color[i], theme_idx)
                    points.append(point)
                if len(points) < 2:
                    continue
                r = {"type": "colorScale", "min": points[0], "max": points[-1]}
                if len(points) > 2:
                    r["mid"] = points[1]
                _with_cf_metadata(r, rule, theme_idx)
                items.append({"cfId": f"cf-{len(items)}", "ranges": ranges, "rule": r})
            elif rule.type == "cellIs":
                formulas = list(rule.formula or [])
                r = {"type": "highlightCell", "operator": rule.operator,
                     "value": formulas[0] if formulas else None}
                if len(formulas) > 1:
                    r["value2"] = formulas[1]
                _with_cf_metadata(r, rule, theme_idx)
                items.append({"cfId": f"cf-{len(items)}", "ranges": ranges,
                              "rule": r})
            elif rule.type == "dataBar" and rule.dataBar and len(rule.dataBar.cfvo) >= 2:
                bar = rule.dataBar
                r = {"type": "dataBar", "min": _cfvo(bar.cfvo[0]), "max": _cfvo(bar.cfvo[-1]),
                     "color": resolve(bar.color, theme_idx)}
                for attr in ("showValue", "minLength", "maxLength"):
                    value = getattr(bar, attr, None)
                    if value is not None:
                        r[attr] = value
                _with_cf_metadata(r, rule, theme_idx)
                items.append({"cfId": f"cf-{len(items)}", "ranges": ranges, "rule": r})
            elif rule.type == "iconSet" and rule.iconSet:
                icons = rule.iconSet
                r = {"type": "iconSet", "iconStyle": icons.iconSet,
                     "thresholds": [_cfvo(obj) for obj in icons.cfvo]}
                for attr in ("showValue", "percent", "reverse"):
                    value = getattr(icons, attr, None)
                    if value is not None:
                        r[attr] = value
                _with_cf_metadata(r, rule, theme_idx)
                items.append({"cfId": f"cf-{len(items)}", "ranges": ranges, "rule": r})
    return items

def read_validation(ws):
    """Map the validation types supported by both converter directions into the Univer resource."""
    supported = {"list", "whole", "decimal", "date", "textLength", "custom"}
    items = []
    for dv in ws.data_validations.dataValidation:
        if dv.type not in supported:
            continue
        item = {"uid": f"dv-{len(items)}", "type": dv.type,
                "allowBlank": bool(dv.allowBlank), "ranges": _range_to_iranges(dv.sqref)}
        if dv.operator is not None:
            item["operator"] = dv.operator
        if dv.formula1 is not None:
            item["formula1"] = dv.formula1
        if dv.formula2 is not None:
            item["formula2"] = dv.formula2
        items.append(item)
    return items

def _store_value(out, value, data_type=None):
    """Store an openpyxl scalar in snapshot form, including JSON-safe date/time caches."""
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        out["v"] = value.isoformat()
        out["t"] = "d"
    elif isinstance(value, dt.timedelta):
        out["v"] = value.total_seconds() / 86400.0
    else:
        out["v"] = value
        if data_type == "e":
            out["t"] = "e"

def _anchor_marker(marker):
    return {"column": marker.col, "columnOffset": marker.colOff,
            "row": marker.row, "rowOffset": marker.rowOff}

def _image_anchor(anchor):
    """Serialize openpyxl's drawing anchors without relying on non-JSON descriptor objects."""
    kind = type(anchor).__name__
    if isinstance(anchor, str):
        return {"type": "cell", "coordinate": anchor}
    if kind == "OneCellAnchor":
        return {"type": "oneCell", "from": _anchor_marker(anchor._from),
                "extent": {"width": anchor.ext.cx, "height": anchor.ext.cy}}
    if kind == "TwoCellAnchor":
        out = {"type": "twoCell", "from": _anchor_marker(anchor._from),
               "to": _anchor_marker(anchor.to)}
        if anchor.editAs is not None:
            out["editAs"] = anchor.editAs
        return out
    if kind == "AbsoluteAnchor":
        return {"type": "absolute", "position": {"x": anchor.pos.x, "y": anchor.pos.y},
                "extent": {"width": anchor.ext.cx, "height": anchor.ext.cy}}
    raise ValueError(f"unsupported image anchor {kind}")

def _read_images(ws):
    """Return JSON-safe embedded images plus per-image errors that must be reported, never hidden."""
    images, errors = [], []
    for index, image in enumerate(ws._images):
        try:
            data = image._data()
            images.append({"format": image.format or "png", "data": base64.b64encode(data).decode("ascii"),
                           "width": image.width, "height": image.height,
                           "anchor": _image_anchor(image.anchor)})
        except Exception as exc:
            errors.append(f"{ws.title} image {index + 1}: {exc}")
    return images, errors

def scan_embedded_objects(path):
    """Count chart/pivot/macro/image/external-link parts in the .xlsx zip — things openpyxl won't
    reconstruct. Runs as a preflight over the raw package so it is fail-safe (never depends on a
    successful cell import)."""
    counts = {"charts": 0, "pivots": 0, "macros": 0, "images": 0, "external_links": 0}
    with zipfile.ZipFile(path) as z:
        for n in z.namelist():
            if n.startswith("xl/charts/") and n.endswith(".xml"):
                counts["charts"] += 1
            elif n.startswith("xl/pivotTables/") and n.endswith(".xml"):
                counts["pivots"] += 1
            elif n == "xl/vbaProject.bin":
                counts["macros"] += 1
            elif n.startswith("xl/media/"):
                counts["images"] += 1
            elif n.startswith("xl/externalLinks/") and n.endswith(".xml"):
                counts["external_links"] += 1
    return counts

def convert(in_path, out_snapshot, out_objects=None):
    # Fail-safe preflight: scan the raw package FIRST and persist the gap inventory, so charts/pivots/
    # macros are reported even if cell import later fails. Detection must never depend on success.
    detected = scan_embedded_objects(in_path)
    preflight = {k: v for k, v in detected.items() if v}
    note = ("Preflight detected before import (reconstruction not yet confirmed): "
            + (", ".join(f"{v} {k}" for k, v in preflight.items()) or "none") + ".")
    if out_objects:
        _atomic_write_json(out_objects, {"charts": [], "pivots": [], "_import_note": note})
    if preflight:
        print("  ⚠ preflight — " + note, file=sys.stderr)

    theme_idx = read_theme(in_path)   # workbook theme palette, for resolving theme/indexed colours
    wb = load_workbook(in_path, data_only=False)
    cached_wb = load_workbook(in_path, data_only=True)
    styles = {}
    key_to_id = {}
    seq = [0]
    def intern(s):
        if not s:
            return None
        key = json.dumps(s, sort_keys=True)
        if key in key_to_id:
            return key_to_id[key]
        sid = f"s{seq[0]}"; seq[0] += 1
        styles[sid] = s; key_to_id[key] = sid
        return sid

    sheet_order, sheets = [], {}
    cf_by_sheet = {}
    dv_by_sheet = {}
    captured_images = 0
    image_errors = []
    for i, ws in enumerate(wb.worksheets):
        sid = f"sheet-{i + 1}"
        sheet_order.append(sid)
        cf_items = read_cond_format(ws, theme_idx)
        if cf_items:
            cf_by_sheet[sid] = cf_items
        dv_items = read_validation(ws)
        if dv_items:
            dv_by_sheet[sid] = dv_items
        cached_ws = cached_wb[ws.title]
        cell_data = {}
        for row in ws.iter_rows():
            for cell in row:
                if (cell.value is None and not cell.has_style and
                        getattr(cell, "comment", None) is None and getattr(cell, "hyperlink", None) is None):
                    continue
                out = {}
                v = cell.value
                if cell.data_type == "f" and isinstance(v, str):
                    out["f"] = v if v.startswith("=") else "=" + v
                    cached_cell = cached_ws[cell.coordinate]
                    _store_value(out, cached_cell.value, cached_cell.data_type)
                elif ArrayFormula and isinstance(v, (ArrayFormula, DataTableFormula)):
                    txt = getattr(v, "text", None) or ""   # keep the formula text; array/spill semantics not modeled
                    out["f"] = txt if txt.startswith("=") else "=" + txt
                    cached_cell = cached_ws[cell.coordinate]
                    _store_value(out, cached_cell.value, cached_cell.data_type)
                elif v is not None:
                    _store_value(out, v, cell.data_type)
                    if isinstance(v, str) and v.startswith("=") and cell.data_type != "f":
                        out["t"] = "s"   # literal text starting with '=' — keep it out of formula-land on export
                sd = intern(style_from(cell, theme_idx))
                if sd:
                    out["s"] = sd
                if cell.comment is not None:
                    out["comment"] = {"text": cell.comment.text, "author": cell.comment.author}
                if cell.hyperlink is not None:
                    link = {"target": cell.hyperlink.target or cell.hyperlink.location}
                    if cell.hyperlink.location is not None:
                        link["location"] = cell.hyperlink.location
                    if cell.hyperlink.display is not None:
                        link["display"] = cell.hyperlink.display
                    if cell.hyperlink.tooltip is not None:
                        link["tooltip"] = cell.hyperlink.tooltip
                    out["hyperlink"] = link
                if out:
                    cell_data.setdefault(str(cell.row - 1), {})[str(cell.column - 1)] = out
        merges = [{"startRow": r.min_row - 1, "endRow": r.max_row - 1,
                   "startColumn": r.min_col - 1, "endColumn": r.max_col - 1} for r in ws.merged_cells.ranges]
        column_data = {}
        column_dimensions = []
        for letter, dim in ws.column_dimensions.items():
            index = column_index_from_string(letter)
            start = (dim.min or index) - 1
            end = (dim.max or index) - 1
            native = {"startColumn": start, "endColumn": end}
            browser = {}
            if dim.width is not None:
                native["width"] = dim.width
                browser["w"] = dim.width * 7
            if dim.hidden:
                native["hidden"] = True
                browser["hd"] = 1
            if dim.outlineLevel:
                native["outlineLevel"] = dim.outlineLevel
            if dim.collapsed:
                native["collapsed"] = True
            if dim.bestFit:
                native["bestFit"] = True
            column_dimensions.append(native)
            if browser:
                column_data[str(start)] = browser
        row_data = {}
        for rnum, dim in ws.row_dimensions.items():
            native = {}
            if dim.height is not None:
                native["height"] = dim.height
                native["h"] = dim.height / 0.75
            if dim.hidden:
                native["hd"] = 1
            if dim.outlineLevel:
                native["outlineLevel"] = dim.outlineLevel
            if dim.collapsed:
                native["collapsed"] = True
            if native:
                row_data[str(rnum - 1)] = native
        images, errors = _read_images(ws)
        captured_images += len(images)
        image_errors.extend(errors)
        freeze = None
        if ws.freeze_panes:
            m = re.match(r"([A-Z]+)(\d+)", ws.freeze_panes)
            if m:
                col = column_index_from_string(m.group(1)) - 1
                rown = int(m.group(2)) - 1
                freeze = {"xSplit": col, "ySplit": rown, "startRow": rown, "startColumn": col}
        # JSON-safe converter extensions: sheetState; exact columnDimensions spans; rowData.height;
        # cell comment/hyperlink records; and base64-encoded per-sheet images with drawing anchors.
        sheet = {"id": sid, "name": ws.title, "sheetState": ws.sheet_state,
                 "rowCount": max(ws.max_row + 10, 50),
                 "columnCount": max(ws.max_column + 5, 26), "cellData": cell_data, "mergeData": merges,
                 "columnData": column_data, "columnDimensions": column_dimensions,
                 "rowData": row_data, "images": images}
        if freeze:
            sheet["freeze"] = freeze
        sheets[sid] = sheet

    resources = []
    dn = {}
    definitions = list(wb.defined_names.values())
    for ws in wb.worksheets:
        definitions.extend(ws.defined_names.values())
    for j, defn in enumerate(definitions):
        item = {"id": f"dn-{j}", "name": defn.name, "formulaOrRefString": defn.value}
        if defn.localSheetId is not None:
            item["localSheetId"] = defn.localSheetId
        dn[f"dn-{j}"] = item
    if dn:
        resources.append({"name": "SHEET_DEFINED_NAME_PLUGIN", "data": json.dumps(dn)})
    if cf_by_sheet:
        resources.append({"name": "SHEET_CONDITIONAL_FORMATTING_PLUGIN", "data": json.dumps(cf_by_sheet)})
    if dv_by_sheet:
        resources.append({"name": "SHEET_DATA_VALIDATION_PLUGIN", "data": json.dumps(dv_by_sheet)})

    snapshot = {"id": "book-imported", "name": (wb.properties.title or "Imported"),
                "appVersion": "0.10.2", "locale": "enUS",
                "sheetOrder": sheet_order, "styles": styles, "sheets": sheets, "resources": resources}
    _atomic_write_json(out_snapshot, snapshot)   # temp+rename: no truncated file on a mid-write crash
    unresolved_images = max(detected.get("images", 0) - captured_images, len(image_errors))
    unresolved = {k: v for k, v in detected.items() if v and k != "images"}
    if unresolved_images:
        unresolved["images"] = unresolved_images
    note = ("Detected in the source but NOT reconstructed by import: "
            + (", ".join(f"{v} {k}" for k, v in unresolved.items()) or "none")
            + ". Charts/pivots must be re-declared; macros/external-links remain out of scope.")
    if out_objects:
        _atomic_write_json(out_objects, {"charts": [], "pivots": [], "_import_note": note})
    if unresolved:
        print("  ⚠ import gap — " + note, file=sys.stderr)
    for error in image_errors:
        print("  ⚠ image not reconstructed — " + error, file=sys.stderr)
    print(f"wrote {out_snapshot} ({len(sheet_order)} sheet(s), {len(styles)} styles)")

def main():
    if len(sys.argv) < 3:
        print("usage: xlsx-to-snapshot.py <in.xlsx> <out.snapshot.json> [out.objects.json]", file=sys.stderr)
        sys.exit(2)
    convert(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else None)

if __name__ == "__main__":
    main()
