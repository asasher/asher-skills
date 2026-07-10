import base64
import copy
import io
import json
import os
import runpy
import subprocess
import sys
import tempfile
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from validate_objects import validate_against_snapshot, validate_objects

wb = load_workbook(sys.argv[1])
cached_wb = load_workbook(sys.argv[1], data_only=True)
ws = wb["Sales"]
checks = []
def ok(name, cond, got=""):
    checks.append((name, bool(cond), got))

ok("title text", ws["A1"].value == "Regional Sales", ws["A1"].value)
ok("title bold", ws["A1"].font.bold is True)
ok("title fill", ws["A1"].fill.fgColor.rgb == "FF1F4E79", ws["A1"].fill.fgColor.rgb)
ok("title center", ws["A1"].alignment.horizontal == "center")
ok("merge A1:D1", "A1:D1" in [str(r) for r in ws.merged_cells.ranges], [str(r) for r in ws.merged_cells.ranges])
ok("header bottom border", ws["A2"].border.bottom.style == "thin", ws["A2"].border.bottom.style)
ok("currency numfmt", ws["D3"].number_format == "$#,##0.00", ws["D3"].number_format)
ok("total formula", ws["D7"].value == "=SUM(D3:D6)", ws["D7"].value)
ok("cached formula value", cached_wb["Sales"]["D7"].value == 4200, cached_wb["Sales"]["D7"].value)
ok("col A width", ws.column_dimensions["A"].width and ws.column_dimensions["A"].width > 15, ws.column_dimensions["A"].width)
ok("row1 height", ws.row_dimensions[1].height == 21, ws.row_dimensions[1].height)
ok("freeze A3", ws.freeze_panes == "A3", ws.freeze_panes)
ok("data validation list", any(dv.type == "list" for dv in ws.data_validations.dataValidation), [dv.type for dv in ws.data_validations.dataValidation])
cf_types = [rule.type for cf in ws.conditional_formatting for rule in cf.rules]
ok("colorScale CF", "colorScale" in cf_types, cf_types)
ok("named range TotalRevenue", "TotalRevenue" in wb.defined_names, list(wb.defined_names))
ok("chart present", len(ws._charts) == 1, len(ws._charts))
ok("chart title", ws._charts[0].title is not None if ws._charts else False)

# pivot (flattened) sheet
pv = wb["Pivot by Region"]
vals = {pv.cell(row=r, column=1).value: pv.cell(row=r, column=2).value for r in range(2, 6)}
ok("pivot North=2800", vals.get("North") == 2800, vals)
ok("pivot South=1400", vals.get("South") == 1400, vals)
ok("pivot Grand=4200", vals.get("Grand Total") == 4200, vals)

# Full import/export regression: cached formula values, validation quoting/types, and exact CFVOs.
converter = Path(__file__).resolve().parent
env = dict(os.environ, PYTHONDONTWRITEBYTECODE="1")
with tempfile.TemporaryDirectory(prefix="spreadsheet-converter-test-", dir="/tmp") as td:
    td = Path(td)
    structured_path = td / "structured-source.xlsx"
    structured = load_workbook(sys.argv[1])
    structured_sales = structured["Sales"]
    structured_sales["A3"].comment = Comment("Structure survives", "Spreadsheet Loop")
    structured_sales["B3"].hyperlink = "https://example.com/widget"
    structured_sales["B3"].hyperlink.display = "Widget docs"
    structured_sales.column_dimensions.group("B", "D", hidden=True, outline_level=2)
    structured_sales.column_dimensions["B"].width = 13.37
    structured_sales.row_dimensions[3].height = 19.35
    structured["Pivot by Region"].sheet_state = "veryHidden"
    pixel = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
    )
    image = Image(io.BytesIO(pixel))
    image.width, image.height = 24, 18
    structured_sales.add_image(image, "E10")
    structured.save(structured_path)
    sample_snapshot = json.load(open(converter / "sample.snapshot.json"))
    runpy.run_path(str(converter / "snapshot-to-xlsx.py"))["_inject_formula_caches"](
        str(structured_path), sample_snapshot, structured.epoch
    )

    imported_path = td / "imported.snapshot.json"
    objects_path = td / "imported.objects.json"
    subprocess.run([sys.executable, str(converter / "xlsx-to-snapshot.py"), str(structured_path),
                    str(imported_path), str(objects_path)], check=True, capture_output=True, text=True, env=env)
    imported = json.load(open(imported_path))

    def resource(snapshot, name):
        entry = next(r for r in snapshot["resources"] if r["name"] == name)
        return entry, json.loads(entry["data"]) if isinstance(entry["data"], str) else entry["data"]

    formula_cell = imported["sheets"]["sheet-1"]["cellData"]["6"]["3"]
    ok("cached formula imported", formula_cell.get("f") == "=SUM(D3:D6)" and formula_cell.get("v") == 4200,
       formula_cell)

    sales_sheet = imported["sheets"]["sheet-1"]
    pivot_sheet = imported["sheets"]["sheet-2"]
    span = next((d for d in sales_sheet.get("columnDimensions", [])
                 if d.get("startColumn") == 1 and d.get("endColumn") == 3), None)
    ok("sheet visibility imported", sales_sheet.get("sheetState") == "visible" and
       pivot_sheet.get("sheetState") == "veryHidden",
       (sales_sheet.get("sheetState"), pivot_sheet.get("sheetState")))
    ok("column span and exact width imported", span is not None and span.get("width") == 13.37 and
       span.get("hidden") is True and span.get("outlineLevel") == 2, span)
    ok("exact row height imported", sales_sheet["rowData"]["2"].get("height") == 19.35,
       sales_sheet["rowData"]["2"])
    ok("comment imported", sales_sheet["cellData"]["2"]["0"].get("comment") ==
       {"text": "Structure survives", "author": "Spreadsheet Loop"},
       sales_sheet["cellData"]["2"]["0"].get("comment"))
    imported_link = sales_sheet["cellData"]["2"]["1"].get("hyperlink")
    ok("hyperlink imported", imported_link.get("target") == "https://example.com/widget" and
       imported_link.get("display") == "Widget docs", imported_link)
    imported_images = sales_sheet.get("images", [])
    ok("image imported", len(imported_images) == 1 and imported_images[0].get("format") == "png" and
       imported_images[0].get("data") == base64.b64encode(pixel).decode("ascii") and
       imported_images[0].get("anchor", {}).get("type") == "oneCell" and
       imported_images[0]["anchor"].get("extent") == {"width": 228600, "height": 171450}, imported_images)

    dv_entry, validations = resource(imported, "SHEET_DATA_VALIDATION_PLUGIN")
    imported_dv = validations["sheet-1"][0]
    ok("validation imported", imported_dv.get("type") == "list" and
       imported_dv.get("formula1") == '"Widget,Gadget,Gizmo"' and imported_dv.get("allowBlank") is True,
       imported_dv)
    validations["sheet-1"].extend([
        {"uid": "dv-named", "type": "list", "formula1": "TotalRevenue", "allowBlank": False,
         "ranges": [{"startRow": 2, "endRow": 5, "startColumn": 4, "endColumn": 4}]},
        {"uid": "dv-decimal", "type": "decimal", "operator": "between", "formula1": 0,
         "formula2": 5000, "allowBlank": False,
         "ranges": [{"startRow": 2, "endRow": 5, "startColumn": 3, "endColumn": 3}]},
        {"uid": "dv-whole", "type": "whole", "operator": "between", "formula1": 1,
         "formula2": 10, "allowBlank": False,
         "ranges": [{"startRow": 2, "endRow": 5, "startColumn": 5, "endColumn": 5}]},
        {"uid": "dv-date", "type": "date", "operator": "between", "formula1": "DATE(2026,1,1)",
         "formula2": "DATE(2026,12,31)", "allowBlank": True,
         "ranges": [{"startRow": 2, "endRow": 5, "startColumn": 6, "endColumn": 6}]},
        {"uid": "dv-text", "type": "textLength", "operator": "lessThanOrEqual", "formula1": 20,
         "allowBlank": True,
         "ranges": [{"startRow": 2, "endRow": 5, "startColumn": 7, "endColumn": 7}]},
        {"uid": "dv-custom", "type": "custom", "formula1": "=LEN(I3)>0", "allowBlank": False,
         "ranges": [{"startRow": 2, "endRow": 5, "startColumn": 8, "endColumn": 8}]},
    ])
    dv_entry["data"] = json.dumps(validations)

    cf_entry, conditional = resource(imported, "SHEET_CONDITIONAL_FORMATTING_PLUGIN")
    imported_rule = conditional["sheet-1"][0]["rule"]
    imported_scale = sorted(imported_rule["config"], key=lambda point: point["index"])
    ok("CF thresholds imported", [point["value"].get("type") for point in imported_scale] ==
       ["min", "percentile", "max"] and imported_scale[1]["value"].get("value") == 50,
       imported_rule)
    imported_rule.update({"priority": 7, "stopIfTrue": True,
                          "style": {"bl": 1, "it": 1, "fs": 9, "ff": "Arial",
                                    "cl": {"rgb": "#123456"}}})
    imported_scale[0]["value"].update({"type": "num", "value": 1000})
    imported_scale[1]["value"].update({"type": "percentile", "value": 37})
    imported_scale[2]["value"].update({"type": "num", "value": 2000})
    cf_entry["data"] = json.dumps(conditional)

    objects = json.load(open(objects_path))
    objects["charts"] = [{
        "id": "chart-stable", "sheet": "Stale Sales Name", "sheetId": "sheet-1", "type": "bar",
        "title": "Revenue by row", "categories": "Sales!B3:B6", "values": ["Sales!D3:D6"],
        "seriesTitles": ["Revenue"], "anchor": "F3"
    }]
    objects["pivots"] = [{
        "id": "pivot-stable", "sheet": "Stale Pivot Name", "sheetId": "sheet-2",
        "source": "Sales!A2:D6", "rows": ["Region"],
        "values": [{"field": "Revenue", "agg": "sum"}], "anchor": "A1"
    }]
    object_errors = validate_objects(objects) + validate_against_snapshot(objects, imported)
    ok("stable sheet IDs validate", not object_errors, object_errors)
    invalid_objects = copy.deepcopy(objects)
    invalid_objects["charts"][0]["sheetId"] = "sheet-missing"
    invalid_errors = validate_against_snapshot(invalid_objects, imported)
    ok("invalid sheet ID rejected", any("sheet-missing" in e for e in invalid_errors), invalid_errors)
    json.dump(objects, open(objects_path, "w"), indent=2)

    mutated_path = td / "mutated.snapshot.json"
    json.dump(imported, open(mutated_path, "w"), indent=2)
    roundtrip_path = td / "roundtrip.xlsx"
    subprocess.run([sys.executable, str(converter / "snapshot-to-xlsx.py"), str(mutated_path),
                    str(roundtrip_path), str(objects_path)], check=True, capture_output=True, text=True, env=env)
    roundtrip = load_workbook(roundtrip_path)
    roundtrip_cached = load_workbook(roundtrip_path, data_only=True)
    rt_dv = list(roundtrip["Sales"].data_validations.dataValidation)
    ok("inline validation quoted", any(d.type == "list" and d.formula1 == '"Widget,Gadget,Gizmo"'
                                        for d in rt_dv), [(d.type, d.formula1) for d in rt_dv])
    ok("named validation unquoted", any(d.type == "list" and d.formula1 == "TotalRevenue" for d in rt_dv),
       [(d.type, d.formula1) for d in rt_dv])
    ok("decimal validation round-trip", any(d.type == "decimal" and d.operator == "between" and
                                              d.formula1 == "0" and d.formula2 == "5000" for d in rt_dv),
       [(d.type, d.operator, d.formula1, d.formula2) for d in rt_dv])
    ok("validation types round-trip", {"list", "whole", "decimal", "date", "textLength", "custom"} <=
       {d.type for d in rt_dv}, [(d.type, d.operator, d.formula1, d.formula2) for d in rt_dv])
    ok("cached formula round-trip", roundtrip_cached["Sales"]["D7"].value == 4200,
       roundtrip_cached["Sales"]["D7"].value)
    rt_sales = roundtrip["Sales"]
    rt_pivot = roundtrip["Pivot by Region"]
    rt_span = rt_sales.column_dimensions["B"]
    ok("sheet visibility round-trip", rt_sales.sheet_state == "visible" and
       rt_pivot.sheet_state == "veryHidden", (rt_sales.sheet_state, rt_pivot.sheet_state))
    ok("column span and exact width round-trip", rt_span.min == 2 and rt_span.max == 4 and
       rt_span.width == 13.37 and rt_span.hidden is True and rt_span.outlineLevel == 2,
       (rt_span.min, rt_span.max, rt_span.width, rt_span.hidden, rt_span.outlineLevel))
    ok("exact row height round-trip", rt_sales.row_dimensions[3].height == 19.35,
       rt_sales.row_dimensions[3].height)
    ok("comment round-trip", rt_sales["A3"].comment is not None and
       rt_sales["A3"].comment.text == "Structure survives" and
       rt_sales["A3"].comment.author == "Spreadsheet Loop", rt_sales["A3"].comment)
    ok("hyperlink round-trip", rt_sales["B3"].hyperlink is not None and
       rt_sales["B3"].hyperlink.target == "https://example.com/widget" and
       rt_sales["B3"].hyperlink.display == "Widget docs", rt_sales["B3"].hyperlink)
    rt_images = rt_sales._images
    ok("image round-trip", len(rt_images) == 1 and
       getattr(rt_images[0].anchor, "_from", None) is not None and
       rt_images[0].anchor._from.col == 4 and rt_images[0].anchor._from.row == 9 and
       rt_images[0].anchor.ext.cx == 228600 and rt_images[0].anchor.ext.cy == 171450, rt_images)
    ok("chart stable ID resolution", len(rt_sales._charts) == 1, len(rt_sales._charts))
    stable_vals = {rt_pivot.cell(row=r, column=1).value: rt_pivot.cell(row=r, column=2).value
                   for r in range(2, 6)}
    ok("pivot stable ID resolution", stable_vals.get("Grand Total") == 4200, stable_vals)

    color_rule = next(rule for cf in roundtrip["Sales"].conditional_formatting for rule in cf.rules
                      if rule.type == "colorScale")
    thresholds = [(obj.type, obj.val) for obj in color_rule.colorScale.cfvo]
    ok("CF thresholds exported", thresholds == [("num", 1000.0), ("percentile", 37.0), ("num", 2000.0)],
       thresholds)
    ok("CF priority and stop", color_rule.priority == 7 and color_rule.stopIfTrue is True,
       (color_rule.priority, color_rule.stopIfTrue))
    ok("CF differential font", color_rule.dxf is not None and color_rule.dxf.font is not None and
       color_rule.dxf.font.bold is True and color_rule.dxf.font.italic is True and
       color_rule.dxf.font.name == "Arial" and color_rule.dxf.font.sz == 9,
       color_rule.dxf.font if color_rule.dxf else None)

    reimported_path = td / "reimported.snapshot.json"
    subprocess.run([sys.executable, str(converter / "xlsx-to-snapshot.py"), str(roundtrip_path),
                    str(reimported_path)], check=True, capture_output=True, text=True, env=env)
    reimported = json.load(open(reimported_path))
    _, reimported_cf = resource(reimported, "SHEET_CONDITIONAL_FORMATTING_PLUGIN")
    reimported_rule = reimported_cf["sheet-1"][0]["rule"]
    reimported_scale = sorted(reimported_rule["config"], key=lambda point: point["index"])
    ok("CF thresholds re-imported", [(point["value"].get("type"), point["value"].get("value"))
                                      for point in reimported_scale] ==
       [("num", 1000.0), ("percentile", 37.0), ("num", 2000.0)], reimported_rule)

npass = sum(1 for _, p, _ in checks if p)
for name, p, got in checks:
    print(f"{'PASS' if p else 'FAIL'}  {name}" + ("" if p else f"   got={got!r}"))
print(f"\n{npass}/{len(checks)} checks passed")
sys.exit(0 if npass == len(checks) else 1)
